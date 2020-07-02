import argparse
import openpyxl


"""
Compare two excell files.
"Compare" means sheets order and all cell in sheets.

todo : 
"""

class WorkBook:
    def __init__(self, name):
        self.workbook = openpyxl.load_workbook(name)
        self.init_bookinfo(name)

    def init_bookinfo(self, name):
        self.book_info = BookInfo(name)
        self.book_info.sheetnames = self.workbook.sheetnames


class BookInfo:
    def __init__(self, name):
        self.name = name
        self.sheetnames = []
        self.tmp_sheetname = ""


def get_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook_path_src")
    parser.add_argument("workbook_path_dst")
    return parser.parse_args()


def output_log(message):
    print(
        f"***************************\n" \
        f"{message}\n" \
        f"***************************\n"
    )


def makeup_for_sheets(shortage, complete):
    if len(shortage.sheetnames) < len(complete.sheetnames):
        shortage_num = len(complete.sheetnames) - len(shortage.sheetnames)
        shortage.sheetnames += [None] * shortage_num
   

def compare_order(src, dst):
    for i, (src_sheetname, dst_sheetname) in enumerate(zip(src.sheetnames, dst.sheetnames)):
        if src_sheetname != dst_sheetname:
            output_log(
                f"In {src.name}, {i+1}th Sheet is {src_sheetname}.\n" \
                f"But In {dst.name}, {i+1}th Sheet is {dst_sheetname}."
            )


def out_data(book_info, cell, is_none):
    if is_none:
        value = "None"
    else:
        value = cell.value
    return f"FileName : {book_info.name}, " \
            f"SheetName : {book_info.tmp_sheetname}, " \
            f"Value : {value}, " \
            f"Cordinate : {cell.coordinate}"


def compare_lines(src, dst, book_info):
    cells_count = max(len(src), len(dst))
    for i in range(cells_count):
        if len(src) < i + 1:
            if not dst[i].value is None:
                output_log(
                    out_data(book_info["src"], dst[i], True) + "\n"
                    + out_data(book_info["dst"], dst[i], False)
                )
        elif len(dst) < i + 1:
            if not src[i].value is None:
                output_log(
                    out_data(book_info["src"], src[i], False) + "\n"
                    + out_data(book_info["dst"], src[i], True)
                )
        elif src[i].value != dst[i].value:
            output_log(
                out_data(book_info["src"], src[i], False) + "\n"
                + out_data(book_info["dst"], dst[i], False)
            )


def compare_sheets(src, dst, book_info):
    lines_count = max(src.max_column, dst.max_column)
    for i in range(1, lines_count + 1):
        compare_lines(src[i], dst[i], book_info)


def select_sheets_for_compare(src, dst):
    book_info = {"src":src.book_info, "dst":dst.book_info}
    for i in range(len(src.book_info.sheetnames)):
        src.book_info.tmp_sheetname = src.book_info.sheetnames[i]
        dst.book_info.tmp_sheetname = dst.book_info.sheetnames[i]
        if src.book_info.tmp_sheetname is None or dst.book_info.tmp_sheetname is None:
            continue
        else:
            compare_sheets(src.workbook[src.book_info.tmp_sheetname],
                            dst.workbook[dst.book_info.tmp_sheetname], book_info)
    # todo : シート名で比較

def main():
    # パラメータの取得
    args = get_args()
    # 入力ファイルの情報取得
    src = WorkBook(args.workbook_path_src)
    dst = WorkBook(args.workbook_path_dst)
    # シート数を同数にする
    makeup_for_sheets(src.book_info, dst.book_info)
    makeup_for_sheets(dst.book_info, src.book_info)
    # シートの並び順を比較
    compare_order(src.book_info, dst.book_info)
    # シートの内容を比較
    select_sheets_for_compare(src, dst)


if __name__ == "__main__":
    main()